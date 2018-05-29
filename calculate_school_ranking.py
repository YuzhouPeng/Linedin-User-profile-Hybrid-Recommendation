from openpyxl import load_workbook
import pandas as pd
import globalparameter, re

def load_university_datafile():


    wb = load_workbook(
        filename=r'/Users/pengyuzhou/Google Drive/Linkedin_datafile/2018_qs_world_university_rankings.xlsx')
    sheets = wb.get_sheet_names()  # 获取所有表格(worksheet)的名字
    sheet0 = sheets[0]  # 第一个表格的名称
    ws = wb.get_sheet_by_name('2018QWUR')  # 获取特定的 worksheet
    # 获取表格所有行和列，两者都是可迭代的
    rows = ws.rows
    columns = ws.columns

    university_name = []
    university_ranking = []
    university_index = []
    index = 0
    # 通过坐标读取值
    # print (ws.cell('B12').value)  # B 表示列，12 表示行

    for r in range(2, 102):
        university_name.append((ws.cell(row=r, column=3).value).lower())
        university_index.append(index)
        index = index + 1
        university_ranking.append(5)
    for r in range(102, 202):
        university_name.append((ws.cell(row=r, column=3).value).lower())
        university_index.append(index)
        index = index + 1
        university_ranking.append(4)
    for r in range(202, 302):
        university_name.append((ws.cell(row=r, column=3).value).lower())
        university_index.append(index)
        index = index + 1
        university_ranking.append(3)
    for r in range(302, 402):
        university_name.append((ws.cell(row=r, column=3).value).lower())
        university_index.append(index)
        index = index + 1
        university_ranking.append(2)
    for r in range(402, 502):
        university_name.append((ws.cell(row=r, column=3).value).lower())
        university_index.append(index)
        index = index + 1
        university_ranking.append(1)

    df = pd.DataFrame(
        {'index': university_index, 'univerisity_name': university_name, 'univerisity_ranking': university_ranking})
    df.to_csv('/Users/pengyuzhou/Google Drive/Linkedin_datafile/university_ranking.csv')

def calculate_user_school_ranking(folderpath, job_title_name,job_title_data_name,extract_name):
    return